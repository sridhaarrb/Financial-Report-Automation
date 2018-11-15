import pandas as pd
import numpy as np
import openpyxl
import os
Account1 = raw_input('Enter the Account: ')
Rep_Month = input('Enter the Month for which Report is done: ')
home = os.path.expanduser('~')


wb = openpyxl.load_workbook(home+'\Monthly Financial Report.xlsx')
ws = wb.active
df_OBcado = pd.read_excel(home+'\OBcado2.xlsx',encoding="utf-8-sig")
df_OBcado.dropna(how='all',axis=0,inplace = True)
df_OBcado.dropna(how='all',axis=1,inplace = True)
y = [i.encode('utf-8') for i in df_OBcado.columns]
strip_col = [i.strip() for i in y]
df_OBcado.columns = strip_col
df = df_OBcado.rename(columns={'Created on':'Created_on'})
df['Hours'] = df.Hours.str.encode('utf-8')
df['Hours'] = df.Hours.str.strip()
df['Hours'] = df.Hours.str.replace(',','.')
df['Hours'] = df.Hours.astype('float')
df['Network'] = df.Network.astype('int64')
df['Month'] = df.Date.str.slice(3, 5)
df['Month'] = df.Month.str.encode('utf-8')
df_pivot = df.pivot_table(index='Network',columns='Month', values='Hours',aggfunc=np.sum)
df_pivot.rename(columns = {'01':'1','02':'2','03':'3','04':'4','05':'5','06':'6','07':'7','08':'8','09':'9',
                           '10':'10','11':'11','12':'12'}, inplace = True)
z=0
for i in df_pivot.columns:
   if (int(i) <=Rep_Month):
    z+=1

df_pivot.drop(df_pivot.columns[z:], axis=1,inplace=True)


df_Budget_HRS_Total = pd.read_excel(home+'\Budget.xlsx',sheetname='Hours')
x = [i.encode('utf-8') for i in df_Budget_HRS_Total.columns]
df_Budget_HRS_Total.rename(columns = lambda x: x[0:3], inplace = True)
df_Budget_HRS_Total.rename(columns = {'Acc' : 'Account','Ser' : 'Service Function','Net' : 'Network ID','Jan':'1',
                            'Feb':'2','Mar':'3','Apr':'4','May':'5','Jun':'6','Jul':'7','Aug':'8','Sep':'9',
                           'Oct':'10','Nov':'11','Dec':'12','Tot':'Total','Flo':'Flow'}, inplace = True)

fil_col=[]
for i in range(1,Rep_Month+1):
    fil_col.append(str(i))

df_Budget_HRS_Total.replace('Total',np.nan, inplace= True)
df_Budget_HRS_Total.dropna(axis=0,inplace=True)
col=list(df_Budget_HRS_Total.columns)[2:15]
df_Budget_HRS_Total[col] = df_Budget_HRS_Total[col].apply(pd.to_numeric, errors='coerce', axis=0)
df_Budget_HRS_Total['YTD_Budget'] = df_Budget_HRS_Total[fil_col].sum(axis=1)
df_Budget_HRS = df_Budget_HRS_Total[['Account','Service Function','Network ID','YTD_Budget','Flow']]
df_Account = df_Budget_HRS[df_Budget_HRS['Account']==Account1]
df_Account.set_index('Network ID',inplace=True)
df_Budget_HRS.set_index('Service Function',inplace=True)
df_Final = pd.concat([df_Account,df_pivot],axis=1, join_axes=[df_Account.index])
df_Final.fillna(0, inplace=True)
df_Final.reset_index(inplace=True)
df_Final.drop('Account',inplace=True,axis=1)
df_Final['YTD_Actuals'] = df_Final[list((df_Final.columns[3:]))].sum(axis=1)
df_Final['Variance_Hrs'] = df_Final['YTD_Budget']-df_Final['YTD_Actuals']
df_Final['Variance_%'] = df_Final['Variance_Hrs']/df_Final['YTD_Budget']
a=df_Final.shape[0]+1
df_Final.loc[a,'Service Function'] = 'Overall'
cols = list(df_Final.columns)
cols.pop(cols.index('Flow'))
df_Final = df_Final[cols+['Flow']]
for m in range(2,len(df_Final.columns)-2):
    col_name = df_Final.columns[m]
    df_Final.loc[a,df_Final.columns[m]] = df_Final[col_name].sum()


df_Final.loc[a,'Variance_%'] = df_Final.loc[a,'Variance_Hrs']/df_Final.loc[a,'YTD_Budget']
df_Final.set_index('Service Function',inplace=True)

ideal_col = list(range(1,Rep_Month+1))
index_loc={'1':2,'2':3,'3':4,'4':5,'5':6,'6':7,'7':8,'8':9,'9':10,'10':11,'11':12,'12':13}
for i in ideal_col:
    if str(i) not in list(df_Final.columns[2:Rep_Month+2]):
        df_Final.insert(loc=index_loc.get(str(i)), column=str(i), value=0.00)
        
y = [str(i) for i in df_Final.columns]
df_Final.columns = y
df_Final_Flow=df_Final.copy()
df_Final.drop(['Flow'],inplace=True,axis=1)



# ### Time Booking-BWH ###


df_cji = pd.read_excel(home+'\CJI3.xlsx')
filter=['Time worked direct hours ICRRB', 'Time worked direct hours', 'ICRRB reposting GSC', 'External CS, External Labor']
df_cji.dropna(axis=0, subset=['Period','Fiscal Year'],inplace=True)
df_cji['Period']=df_cji.Period.astype('int64')
df_cji_filter = df_cji.loc[df_cji['Cost element descr.'].isin(filter)]
df_cji_pivot = df_cji_filter.pivot_table(index='Order',columns='Period', values='Total Quantity',aggfunc=np.sum)
z=0
for i in df_cji_pivot.columns:
   if (int(i) <=Rep_Month):
    z+=1

df_cji_pivot.drop(df_cji_pivot.columns[z:], axis=1,inplace=True)
df_Final_BWH = pd.concat([df_Account,df_cji_pivot],axis=1, join_axes=[df_Account.index])


mod_col = [str(i) for i in df_Final_BWH.columns]
df_Final_BWH.columns = mod_col
df_Final_BWH.fillna(0, inplace=True)
df_Final_BWH.reset_index(inplace=True)
df_Final_BWH.drop('Account',inplace=True,axis=1)
df_Final_BWH['YTD_Actuals'] = df_Final_BWH[list((df_Final_BWH.columns[3:]))].sum(axis=1)
df_Final_BWH['Variance_Hrs'] = df_Final_BWH['YTD_Budget']-df_Final_BWH['YTD_Actuals']
df_Final_BWH['Variance_%'] = df_Final_BWH['Variance_Hrs']/df_Final_BWH['YTD_Budget']
b = df_Final_BWH.shape[0]+1
df_Final_BWH.loc[b,'Service Function'] = 'Overall'
cols = list(df_Final_BWH.columns)
cols.pop(cols.index('Flow'))
df_Final_BWH = df_Final_BWH[cols+['Flow']]
for m in range(2,len(df_Final_BWH.columns)-2):
    col_name = df_Final_BWH.columns[m]
    df_Final_BWH.loc[b,df_Final_BWH.columns[m]] = df_Final_BWH[col_name].sum()
    

df_Final_BWH.loc[a,'Variance_%'] = df_Final_BWH.loc[a,'Variance_Hrs']/df_Final_BWH.loc[a,'YTD_Budget']
df_Final_BWH.set_index('Service Function',inplace=True)

ideal_col = list(range(1,Rep_Month+1))
index_loc={'1':2,'2':3,'3':4,'4':5,'5':6,'6':7,'7':8,'8':9,'9':10,'10':11,'11':12,'12':13}
for i in ideal_col:
    if str(i) not in list(df_Final_BWH.columns[2:Rep_Month+2]):
        df_Final_BWH.insert(loc=index_loc.get(str(i)), column=str(i), value=0.00)
        
y = [str(i) for i in df_Final_BWH.columns]
df_Final_BWH.columns = y
df_Final_BWH_Flow = df_Final_BWH.copy()
df_Final_BWH.drop(['Flow'],inplace=True,axis=1)
df_Final_BWH_Flow.drop(['Overall'],inplace=True)


# ### Finance-Functional breakup ###


df_cji1 = pd.read_excel(home+'\CJI3.xlsx')
filter=['Time worked direct hours ICRRB', 'Time worked direct hours', 'ICRRB reposting GSC', 'External CS, External Labor']
df_cji1.dropna(axis=0, subset=['Period','Fiscal Year'],inplace=True)
df_cji1['Period']=df_cji1.Period.astype('int64')
df_cji1_filter = df_cji1.loc[df_cji1['Cost element descr.'].isin(filter)]
df_cji1_pivot = df_cji1_filter.pivot_table(index='Order',columns='Period', values='Val.in rep.cur.',aggfunc=np.sum)
z=0
for i in df_cji1_pivot.columns:
   if (int(i) <=Rep_Month):
    z+=1

df_cji1_pivot.drop(df_cji1_pivot.columns[z:], axis=1,inplace=True)


df_Budget_USD_Total = pd.read_excel(home+'\Budget.xlsx',sheetname='USD')
x = [i.encode('utf-8') for i in df_Budget_USD_Total.columns]
df_Budget_USD_Total.rename(columns = lambda x: x[0:3], inplace = True)
df_Budget_USD_Total.rename(columns = {'Acc' : 'Account','Ser' : 'Service Function','Net' : 'Network ID','Jan':'1',
                            'Feb':'2','Mar':'3','Apr':'4','May':'5','Jun':'6','Jul':'7','Aug':'8','Sep':'9',
                           'Oct':'10','Nov':'11','Dec':'12','Tot':'Total','Flo':'Flow'}, inplace = True)

fil_col=[]
for i in range(1,Rep_Month+1):
    fil_col.append(str(i))

df_Budget_USD_Total.replace('Total',np.nan, inplace= True)
df_Budget_USD_Total.dropna(axis=0,inplace=True)
col=list(df_Budget_USD_Total.columns)[2:15]
df_Budget_USD_Total[col] = df_Budget_USD_Total[col].apply(pd.to_numeric, errors='coerce', axis=0)
df_Budget_USD_Total['YTD_Budget'] = df_Budget_USD_Total[fil_col].sum(axis=1)
df_Budget_USD = df_Budget_USD_Total[['Account','Service Function','Network ID','YTD_Budget','Flow']]
df_Account_USD = df_Budget_USD[df_Budget_USD['Account']==Account1]
df_Account_USD.set_index('Network ID',inplace=True)


df_Final_Finance = pd.concat([df_Account_USD,df_cji1_pivot],axis=1, join_axes=[df_Account_USD.index])


mod_col = [str(i) for i in df_Final_Finance.columns]
df_Final_Finance.columns = mod_col
df_Final_Finance.fillna(0, inplace=True)
df_Final_Finance.reset_index(inplace=True)
df_Final_Finance.drop('Account',inplace=True,axis=1)
df_Final_Finance['YTD_Actuals'] = df_Final_Finance[list((df_Final_Finance.columns[3:]))].sum(axis=1)
df_Final_Finance['Variance_Hrs'] = df_Final_Finance['YTD_Budget']-df_Final_Finance['YTD_Actuals']
df_Final_Finance['Variance_%'] = df_Final_Finance['Variance_Hrs']/df_Final_Finance['YTD_Budget']
b = df_Final_Finance.shape[0]+1
df_Final_Finance.loc[b,'Service Function'] = 'Overall'
cols = list(df_Final_Finance.columns)
cols.pop(cols.index('Flow'))
df_Final_Finance = df_Final_Finance[cols+['Flow']]

for m in range(2,len(df_Final_Finance.columns)-2):
    col_name = df_Final_Finance.columns[m]
    df_Final_Finance.loc[b,df_Final_Finance.columns[m]] = df_Final_Finance[col_name].sum()

df_Final_Finance.loc[b,'Variance_%'] = df_Final_Finance.loc[b,'Variance_Hrs']/df_Final_Finance.loc[b,'YTD_Budget']
df_Final_Finance.set_index('Service Function',inplace=True)
ideal_col = list(range(1,Rep_Month+1))
index_loc={'1':2,'2':3,'3':4,'4':5,'5':6,'6':7,'7':8,'8':9,'9':10,'10':11,'11':12,'12':13}
for i in ideal_col:
    if str(i) not in list(df_Final_Finance.columns[2:Rep_Month+2]):
        df_Final_Finance.insert(loc=index_loc.get(str(i)), column=str(i), value=0.00)
        
y = [str(i) for i in df_Final_Finance.columns]
df_Final_Finance.columns = y
df_Final_Finance_Flow=df_Final_Finance.copy()
df_Final_Finance.drop(['Flow'],inplace=True,axis=1)
df_Final_Finance_Flow.drop(['Overall'],inplace=True)


# ### Writting on Excel ###


from copy import copy
from openpyxl.styles.borders import Border, Side

def Format(sheetname,Rep_Month,data_frame):
    sheet1 = wb.get_sheet_by_name(sheetname)
    sheet1.unmerge_cells('D2:P2')
    sheet1.delete_cols(4+Rep_Month, amount=16-(4+Rep_Month))
    sheet1.merge_cells(start_row=2, start_column=4, end_row=2, end_column=4+Rep_Month)
    sheet1.merge_cells(start_row=2, start_column=5+Rep_Month, end_row=2, end_column=6+Rep_Month)
    if data_frame.shape[0]<5:
        row_diff = 5-data_frame.shape[0]
        sheet1.delete_rows(4, amount=row_diff)

    if data_frame.shape[0]>5:
        row_diff = data_frame.shape[0]-5
        sheet1.unmerge_cells('A10:B10')
        sheet1.insert_rows(5,row_diff)
        
        for i in range(5,5+row_diff):
            for j in range(1,len(data_frame.columns)+2):
                sheet1.cell(row=i, column=j)._style = copy(sheet1.cell(row=4, column=j)._style)

    L_Index=[]
    L_Rows=[]
    Index_Ref=0
    Series_Ref=0


    for row_index,record in data_frame.iterrows():
        L_Index.append(row_index)
        L_Rows.append(record)


    series_len = len(L_Rows[0])
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

    for i in range(4,data_frame.shape[0]+4):    
        for j in range(1,len(data_frame.columns)+2):
            if j==1:
                sheet1.cell(row=i, column=j).value = L_Index[Index_Ref]
            else:
                sheet1.cell(row=i, column=j).value = L_Rows[Index_Ref][Series_Ref]
                Series_Ref =Series_Ref+1
    
        Index_Ref=Index_Ref+1
        Series_Ref=0
    
    row_count=4
    to_be_del=0
    for h in range(4,2000):
        if (sheet1.cell(row=h, column=1).value)!=None:
            row_count+=1
        else:
            break


    if (sheet1.cell(row=row_count, column=1).value)!='Data Source':
        for i in range(row_count,20):
            if (sheet1.cell(row=i, column=1).value)!='Data Source':
                to_be_del+=1
            else:
                break
    else:
        sheet1.insert_rows(row_count,3)
        sheet1.insert_rows(row_count-1,1)
        sheet1.merge_cells(start_row=row_count, start_column=1, end_row=row_count, end_column=2)



    sheet1.delete_rows(row_count, amount=to_be_del)
    sheet1.insert_rows(row_count,3)
    sheet1.insert_rows(row_count-1,1)
    sheet1.merge_cells(start_row=row_count, start_column=1, end_row=row_count, end_column=2)
    sheet1.cell(row=2, column=6+Rep_Month).border = thin_border
    sheet1.cell(row=row_count, column=2).border = thin_border

    wb.save(home+'\Monthly Financial Report.xlsx')


Format('Time Booking-CADO',Rep_Month,df_Final)
Format('Time Booking-BWH',Rep_Month,df_Final_BWH)
Format('Finance-Functional breakup',Rep_Month,df_Final_Finance)    


### Financial Summary ###


Fin_Sum_Budget_USD = df_Budget_USD_Total[df_Budget_USD_Total['Account']==Account1]
Fin_Sum_Budget_HRS = df_Budget_HRS_Total[df_Budget_HRS_Total['Account']==Account1]

Fin_Sum_Budget_USD.name='Budget_USD'
Fin_Sum_Budget_HRS.name='Budget_HRS'
df_Final_Finance_Flow.name='Actual_USD'
df_Final_BWH_Flow.name='Actual_HRS'

data_frame=[Fin_Sum_Budget_USD,Fin_Sum_Budget_HRS,df_Final_Finance_Flow,df_Final_BWH_Flow]
final_dict_collection=[]
final_df_collection=[]
for k in data_frame:
    Flow_Items = k['Flow'].unique()
    for i in Flow_Items:
        dict_collection={}
        A_filter = k[k['Flow']==i]
        for j in range(1,Rep_Month+1):
            dict1=({j: A_filter[str(j)].sum()})
            dict_collection.update(dict1)
        final_dict_collection.append(dict_collection)
        dict2=((k.name)+'_'+i.encode())
        final_df_collection.append(dict2)
        
    
X=zip(final_df_collection,final_dict_collection)



q,r=1,1
for i in X:
    
    if ('USD' in i[0]):
        if (('USD' in i[0]) and (q==1)):
            Final_USD_Summary = pd.DataFrame(i[1].items(), columns=['Month',i[0]])
            Final_USD_Summary.set_index('Month',inplace=True)
            q+=1
        else:
            Final_USD_Summary[i[0]] = pd.Series(i[1], index=Final_USD_Summary.index)
            
    else:
        if (('HRS' in i[0]) and (r==1)):
            Final_HRS_Summary = pd.DataFrame(i[1].items(), columns=['Month',i[0]])
            Final_HRS_Summary.set_index('Month',inplace=True)
            r+=1
        else:
            Final_HRS_Summary[i[0]] = pd.Series(i[1], index=Final_HRS_Summary.index)

    

for i in Final_USD_Summary.columns:
    if i.startswith('Budget'):
        Final_USD_Summary['Variance_USD_'+i[11:]] = Final_USD_Summary[i] - Final_USD_Summary['Actual_USD_'+i[11:]]
        Final_USD_Summary['Variance_USD_'+i[11:]+'%'] = (Final_USD_Summary[i] - Final_USD_Summary['Actual_USD_'+i[11:]])/Final_USD_Summary[i]

for j in Final_HRS_Summary.columns:
    if j.startswith('Budget'):
        Final_HRS_Summary['Variance_HRS_'+j[11:]] = Final_HRS_Summary[j] - Final_HRS_Summary['Actual_HRS_'+j[11:]]
        Final_HRS_Summary['Variance_HRS_'+j[11:]+'%'] = (Final_HRS_Summary[j] - Final_HRS_Summary['Actual_HRS_'+j[11:]])/Final_HRS_Summary[j]


Rev_col_HRS=[]
k = ['Budget_HRS_','Actual_HRS_','Variance_HRS_','Variance_HRS']
for i in Flow_Items:
    for j in k:
        if k.index(j)!= 3:
            Rev_col_HRS.append(str(j+i))
        else:
            Rev_col_HRS.append(str('Variance_HRS_'+i+'%'))
            
Rev_col_USD=[]
k = ['Budget_USD_','Actual_USD_','Variance_USD_','Variance_USD']
for i in Flow_Items:
    for j in k:
        if k.index(j)!= 3:
            Rev_col_USD.append(str(j+i))
        else:
            Rev_col_USD.append(str('Variance_USD_'+i+'%'))
            
Final_HRS_Summary1 = Final_HRS_Summary.reindex(columns = Rev_col_HRS)
Final_USD_Summary1 = Final_USD_Summary.reindex(columns = Rev_col_USD)


import sets
a1 = [str (i) for i in Flow_Items]
a2 = ['Operate','Optimize','IT&C']
s1 = sets.Set(a1)
s2 = sets.Set(a2)
set_list = list(s2-s1)
df1=[Final_USD_Summary1,Final_HRS_Summary1]

if len(set_list)!=0:
    Final_USD_Summary1.name = 'USD'
    Final_HRS_Summary1.name = 'HRS'
    
    for a in df1:
        
        if a.name == 'USD':
            k = ['Budget_USD_','Actual_USD_','Variance_USD_','Variance_USD_']
        if a.name == 'HRS':
            k = ['Budget_HRS_','Actual_HRS_','Variance_HRS_','Variance_HRS_']
        for i in set_list:
            if i is 'Operate':
                j=0
                for m in range(4):
                    if j!=3:
                        a.insert(loc=m, column=k[j]+i, value=0.00)
                    else:
                        a.insert(loc=m, column=k[j]+i+'%', value=0.00)
                    j+=1
                    
            if i is 'Optimize':
                j=0
                for m in range(4,8):
                    if j!=3:
                        a.insert(loc=m, column=k[j]+i, value=0.00)
                    else:
                        a.insert(loc=m, column=k[j]+i+'%', value=0.00)
                    j+=1
                    
            if '&' in i:
                j=0
                for m in range(8,12):
                    if j!=3:
                        a.insert(loc=m, column=k[j]+i, value=0)
                    else:
                        a.insert(loc=m, column=k[j]+i+'%', value=0)
                    j+=1


Final_HRS_Summary1['Budget_HRS_Overall'] = Final_HRS_Summary1[[i for i in Rev_col_HRS if i.startswith('Budget')]].sum(axis=1)
Final_HRS_Summary1['Actual_HRS_Overall'] = Final_HRS_Summary1[[i for i in Rev_col_HRS if i.startswith('Actual')]].sum(axis=1)
Final_HRS_Summary1['Variance_HRS_Overall'] = Final_HRS_Summary1[[i for i in Rev_col_HRS if i.startswith('Variance') and i.endswith('%')==False]].sum(axis=1)
Final_HRS_Summary1['Variance_HRS_Overall%'] = (Final_HRS_Summary1['Budget_HRS_Overall'] - Final_HRS_Summary1['Actual_HRS_Overall'])/Final_HRS_Summary1['Budget_HRS_Overall']
Final_USD_Summary1['Budget_USD_Overall'] = Final_USD_Summary1[[i for i in Rev_col_USD if i.startswith('Budget')]].sum(axis=1)
Final_USD_Summary1['Actual_USD_Overall'] = Final_USD_Summary1[[i for i in Rev_col_USD if i.startswith('Actual')]].sum(axis=1)
Final_USD_Summary1['Variance_USD_Overall'] = Final_USD_Summary1[[i for i in Rev_col_USD if i.startswith('Variance') and i.endswith('%')==False]].sum(axis=1)
Final_USD_Summary1['Variance_USD_Overall%'] = (Final_USD_Summary1['Budget_USD_Overall'] - Final_USD_Summary1['Actual_USD_Overall'])/Final_USD_Summary1['Budget_USD_Overall']


# ### Writing in Financial Summary ###


from copy import copy
from openpyxl.styles.borders import Border, Side
Final_USD_Summary1.name = 'USD'
Final_HRS_Summary1.name = 'HRS'
Month={1:'January', 2:'February', 3:'March', 4:'April', 5:'May', 6:'June', 7:'July', 8:'August', 
       9:'September', 10:'October', 11:'November', 12:'December'}

def Format(sheetname,Rep_Month,df1):
    sheet1 = wb.get_sheet_by_name(sheetname)
    for data_frame in df1:
        L_Index=[]
        L_Rows=[]
        Index_Ref=0
        Series_Ref=0

        for row_index,record in data_frame.iterrows():
            L_Index.append(row_index)
            L_Rows.append(record)

        series_len = len(L_Rows[0])
        
        if data_frame.name=='USD':
            for i in range(7,data_frame.shape[0]+7):
                for j in range(3,len(data_frame.columns)+3):       
                    sheet1.cell(row=i, column=j).value = L_Rows[Index_Ref][Series_Ref]
                    Series_Ref =Series_Ref+1
                
                Index_Ref=Index_Ref+1
                Series_Ref=0
                
            
            for idx in range(7,19):
                if (sheet1.cell(row=idx,column=3).value==0) or (sheet1.cell(row=idx,column=3).value==None):
                    sheet1.row_dimensions[idx].hidden = True
        
        if data_frame.name=='HRS':
            for i in range(25,data_frame.shape[0]+25):    
                for j in range(3,len(data_frame.columns)+3):       
                    sheet1.cell(row=i, column=j).value = L_Rows[Index_Ref][Series_Ref]
                    Series_Ref =Series_Ref+1
    
                Index_Ref=Index_Ref+1
                Series_Ref=0
            
        
            for idx in range(25,37):
                if (sheet1.cell(row=idx,column=3).value==0) or (sheet1.cell(row=idx,column=3).value==None):
                    sheet1.row_dimensions[idx].hidden = True
        
    for i in set_list:
        if i=='Operate':
            for i in ['C','D','E','F']:
                sheet1.column_dimensions[i].hidden= True
                
        if i=='Optimize':
            for i in ['G','H','I','J']:
                sheet1.column_dimensions[i].hidden= True
                
        if i=='IT&C':
            for i in ['K','L','M','N']:
                sheet1.column_dimensions[i].hidden= True
                
    sheet1.cell(row=3, column=2).value = 'Financial & Booking Summary '+Month.get(Rep_Month)
    wb.save(home+'\Monthly Financial Report.xlsx')
    
Format('Financial Summary',Rep_Month,df1)


print '***********End of Script*******************'
